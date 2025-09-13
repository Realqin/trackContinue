import ast
import configparser
import os
import random

import pymysql
import json
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl import load_workbook
from dateutil import tz

# from process_trajectory_data import save_track_to_json


def read_config(config_file='config.ini'):
    """
    读取配置文件
    :param config_file: 配置文件路径
    :return: 配置对象
    """
    config = configparser.ConfigParser()
    config.read(config_file, encoding='utf-8')
    return config


def get_db_connection():
    """
    获取数据库连接
    :return: 数据库连接对象
    """
    # 读取配置
    config = read_config()
    host = config.get('database', 'host')
    port = config.getint('database', 'port')
    database = config.get('database', 'database')
    table = config.get('database', 'table')

    # 连接数据库
    connection = pymysql.connect(
        host=host,
        port=port,
        user='root',
        password='',  # 实际使用时需要配置密码
        database=database,
        charset='utf8mb4'
    )
    print("连接数据库成功")
    return connection


def execute_query(sql, params=None):
    """
    执行数据库查询的通用函数
    :param sql: SQL查询语句
    :param params: 查询参数
    :return: 查询结果
    """
    connection = get_db_connection()
    result = []
    try:
        with connection.cursor() as cursor:
            # 执行查询
            print("执行查询:", sql, params)
            cursor.execute(sql, params)
            result = cursor.fetchall()
            print(f"查询结果: {result}")
    finally:
        connection.close()

    return result


def fetch_trajectory_data(start_time, end_time):
    """
    从数据库读取轨迹数据
    
    :param start_time: 开始时间
    :param end_time: 结束时间
    :return: 符合条件的数据列表
    """
    # 读取配置
    config = read_config()
    table = config.get('database', 'table')

    # 构造查询SQL
    # 查询某个时间段内duration值大于1h（转换为秒），且displacement/rangeMaxDistance < 3的id数据前100条的数据
    sql = f"""
    SELECT id, longitude, latitude, speed, course, len, lastTm, displacement, rangeMaxDistance, duration
    FROM {table}
    WHERE lastTm BETWEEN %s AND %s
    AND duration > 3600
    AND displacement/rangeMaxDistance < 3
    AND rangeMaxDistance > 1000
    ORDER BY lastTm DESC
    LIMIT 100
    """

    rows = execute_query(sql, (start_time, end_time))

    result = []
    # 格式化数据
    for row in rows:
        # 处理时间戳，数据库中是毫秒时间戳格式
        # last_tm_timestamp = row[6] / 1000  # 转换为秒级时间戳
        # last_tm_formatted = datetime.datetime.fromtimestamp(last_tm_timestamp).strftime('%Y-%m-%d %H:%M:%S')

        item = {
            "longitude": row[1],
            "latitude": row[2],
            "speed": row[3],
            "course": row[4],
            "len": row[5],
            "lastTm": row[6],
            "id": row[0]  # 添加ID字段
        }
        result.append(item)

    return result


def fetch_target_ids(table, start_time, end_time, tests_sample_size):
    """
    从数据库读取符合条件的ID列表
    
    :param start_time: 开始时间
    :param end_time: 结束时间
    :return: ID列表
    """

    # 构造查询SQL
    sql = f"""
    SELECT id,count(*)
    FROM {table}
    WHERE lastdt BETWEEN %s AND %s
      AND duration > 3600000          -- 注意：时间单位是毫秒，3600秒 = 3600000毫秒
      AND rangeMaxDistance > 1000
      AND displacement / rangeMaxDistance < 3
    group by id
    having count(*) > 100
    LIMIT %s
    """

    rows = execute_query(sql, (start_time, end_time, tests_sample_size))
    # 提取ID列表
    target_id_list = [row[0] for row in rows]
    return target_id_list


# 遍历每个id在最近1h内的轨迹点，按lastdt倒序，第一部分，每隔track_extract_time取一个点，取10个点--》空track_gap_range--》继续按间隔取10个点
def extract_track_points(table, ids, end_time, track_extract_time, track_gap_range, track_sample_size):
    """
    从数据库读取符合条件的轨迹点

    :param id: 轨迹ID
    :param start_time: 开始时间
    :param end_time: 结束时间
    :return: 轨迹点列表
    """
    # 构造查询SQL
    sql = f"""
    SELECT id,longitude, latitude, speed, course, len, lastTm,lastdt
    FROM {table}
    WHERE id = %s
    AND lastTm BETWEEN %s AND %s
    ORDER BY lastTm DESC
    """

    #转换为时间戳
    beijing = tz.gettz("Asia/Shanghai")
    dt = datetime.strptime(end_time, "%Y-%m-%d %H:%M:%S").replace(tzinfo=beijing)
    lasttm = int(dt.timestamp() * 1000)


    # 随机一个gap范围值
    track_gap_range = ast.literal_eval(track_gap_range)

    # 转换轨迹间隔时间：分钟
    track_extract_time = int(track_extract_time)

    # 存储一些过程数据，
    #

    case_num = 1


    testcases = []
    for id in ids:
        # 每个目标都去查轨迹点,查询近一个小时的轨迹点
        points = execute_query(sql, (id, lasttm - 3600000, lasttm))

        # 每次随机一个轨迹间隔时间
        gap_range = random.randint(track_gap_range[0], track_gap_range[1])
        print("间隔随机gap_range:", gap_range)

        # flag_num:记录轨迹点数量 ； flag_time查到到的lasttm，方便做比较后抽点 ；
        # disappear_points：消失轨迹点 ； appear_points：出现轨迹点；result ：组装好的一条记录 ;
        # others_msg测试数据记录（id1、id1开始时间/结束时间、轨迹间隔时间、id2、id2开始时间/结束时间,样本标记）
        flag_num = 1
        flag_time = 0
        result = []
        disappear_points = []
        appear_points = []
        others_msg = []
        id1_starttime = None
        id1_endtime = None
        id2_starttime = None
        id2_endtime = None

        # 单独给显示轨迹用的
        ids_points = []
        assemble_data = []


        for point in points:

            # 处理时间戳，数据库中是毫秒时间戳格式
            # last_tm_timestamp = row[6] / 1000  # 转换为秒级时间戳
            # last_tm_formatted = datetime.datetime.fromtimestamp(last_tm_timestamp).strftime('%Y-%m-%d %H:%M:%S')

            lastdt = point[7].strftime("%Y-%m-%d %H:%M:%S")  #查到了值，但是不需要组装到数据里


            item = {
                "id": point[0],
                "longitude": f'{point[1]}',
                "latitude": f'{point[2]}',
                "speed": point[3],
                "course": point[4],
                "len": point[5],
                "lastTm": point[6]
            }

            all_item = {
                "id": point[0],
                "longitude": f'{point[1]}',
                "latitude": f'{point[2]}',
                "speed": point[3],
                "course": point[4],
                "len": point[5],
                "lastTm": point[6],
                "lastdt": point[7].strftime("%Y-%m-%d %H:%M:%S")
            }

            # 保存一份全量的轨迹，方便显示排查
            ids_points.append(all_item)
            print("all_points:", ids_points)


            # 按时间点倒序，先提取生成的列表数据
            if len(appear_points) == 0:
                appear_points.append(item)
                flag_time = item['lastTm']
                print("flag_num:", flag_num)
                id2_endtime = lastdt
                flag_num += 1
            #开始提取生成目标的第二个轨迹点，每隔一个track_extract_time取一个点
            elif (flag_num <= track_sample_size) and ((flag_time - item['lastTm']) > (track_extract_time * 60 * 1000)):
                # aa = flag_time - item['lastTm']
                # print("aa:", aa / 60 / 1000)
                # 时间间隔是否超过track_extract_time，是就取当前点
                appear_points.append(item)
                flag_time = item['lastTm']
                print("flag_num:", flag_num)
                id2_starttime = lastdt
                flag_num += 1
            # 开始存消失目标的轨迹点，必须满足间隔track_gap_range后取点
            elif (len(disappear_points) == 0) and ((
                    flag_time - item['lastTm']) > (gap_range * 60 * 1000)) and (
                    flag_num > track_sample_size):
                #
                # bb = (flag_time - item['lastTm']) / 60 / 1000
                # print("间隔:",bb )
                disappear_points.append(item)
                flag_time = item['lastTm']
                print("flag_num:", flag_num)
                id1_endtime = lastdt
                flag_num += 1
            # 每隔一个track_extract_time取一个点，直到数据提取完
            elif ((track_sample_size + 1) < flag_num <= (track_sample_size * 2)) and (
                    flag_time - item['lastTm'] > (track_extract_time  * 60 * 1000)):
                cc = (flag_time - item['lastTm']) / 60 / 1000
                print("间隔:", cc)
                disappear_points.append(item)
                flag_time = item['lastTm']
                print("flag_num:", flag_num)
                id1_starttime = lastdt
                flag_num += 1
            # 数据提取到了就立刻退出循环
            elif flag_num > (track_sample_size * 2):
                break


        others_msg.append({
            'case_num': case_num,
            'id': id,
            'id1_starttime': f'{id1_starttime}',
            'id1_endtime': f'{id1_endtime}',
            'gap_range': gap_range,
            'id2': id,  # 注意：这里还是 id，如果和上面相同可考虑合并
            'id2_starttime': f'{id2_starttime}',
            'id2_endtime': f'{id2_endtime}',
            'direction': '正'  # 方向：正向
        })

        # 因为是倒着找轨迹点的，现在倒一下顺序
        disappear_points.reverse()
        appear_points.reverse()
        result.extend([[disappear_points, appear_points], others_msg])
        print("result:",result)
        print("\n")

        #加一层过滤，确保出现和消失点的数据量都是对的
        if (len(disappear_points) == track_sample_size) and (len(appear_points) == track_sample_size):
            testcases.append(result)

        ids_points.reverse()
        #组装轨迹显示要用的数据
        assemble_data.extend([[ids_points],[disappear_points,appear_points],others_msg])
        save_track_to_json(assemble_data,case_num,'正样本')

        case_num += 1

    print("testcases:",testcases)
    return testcases


def save_track_to_json(data, case_num=None, sample_type=None):
    """
    处理轨迹数据并保存为JSON格式供前端使用

    Args:
        data: 轨迹数据
        case_num: 用例编号
        sample_type: 样本类型（如"正样本"）
    """

    # 确保tracks目录存在
    if not os.path.exists('tracks'):
        os.makedirs('tracks')

    # 生成文件名
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    if case_num is not None and sample_type is not None:
        filename = f"{sample_type}-{case_num}-{timestamp}.json"
    else:
        filename = f"trajectory_data-{timestamp}.json"

    # 保存为JSON文件到tracks目录
    filepath = os.path.join('tracks', filename)

    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"已保存轨迹数据到 {filepath}")


def write_to_excel(data, excel_file, sheet_name='Sheet1'):
    """
    将数据写入Excel文件
    
    :param data: 要写入的数据
    :param excel_file: Excel文件路径
    :param sheet_name: 工作表名称
    """
    try:
        # 加载现有的Excel文件或创建新的
        try:
            workbook = load_workbook(excel_file)
        except FileNotFoundError:
            workbook = Workbook()

        # 选择或创建工作表
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.create_sheet(sheet_name)

        # 设置所有列为文本格式
        from openpyxl.styles import numbers
        for column in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']:
            for row in range(1, len(data) + 2):  # +2 包括标题行和索引从1开始
                worksheet[f'{column}{row}'].number_format = numbers.FORMAT_TEXT

        # 写入表头
        headers = ["测试组", "id1", "id1开始时间", "id1结束时间", "轨迹间隔时间（分钟）", "id2", "id2开始时间", "id2结束时间", "样本值", "程序预测结果", "测试数据"]

        # headers = ["longitude", "latitude", "speed", "course", "len", "lastTm", "data_json", "id"]
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)

        # 写入数据
        for row, item in enumerate(data, 2):
            print("row:", row)

            # 通过在值前添加单引号来强制设置为文本格式
            worksheet.cell(row=row, column=1,value=item[1][0]['case_num'])
            worksheet.cell(row=row, column=2, value=f"{item[1][0]['id']}")
            worksheet.cell(row=row, column=3, value=item[1][0]['id1_starttime'])
            worksheet.cell(row=row, column=4, value=item[1][0]['id1_endtime'])
            worksheet.cell(row=row, column=5, value=item[1][0]['gap_range'])
            worksheet.cell(row=row, column=6, value=f"{item[1][0]['id2']}")
            worksheet.cell(row=row, column=7, value=item[1][0]['id2_starttime'])
            worksheet.cell(row=row, column=8, value=item[1][0]['id2_endtime'])
            worksheet.cell(row=row, column=9, value=item[1][0]['direction'])
            worksheet.cell(row=row, column=11, value=f"{item[0]}")

        # 保存文件
        workbook.save(excel_file)
        print(f"数据已成功写入 {excel_file}")

    except Exception as e:
        print(f"写入Excel时出错: {e}")


def main():
    """
    提取参数，调用函数
    """
    # 示例时间范围，实际使用时请替换为需要的时间范围
    config = read_config()
    table = config.get('database', 'table')
    tests_sample_size = int(config.get('params', 'tests_sample_size'))
    track_sample_size = int(config.get('params', 'track_sample_size'))
    start_time = config.get('params', 'start_time')
    end_time = config.get('params', 'end_time')
    track_extract_time = config.get('params', 'track_extract_time')
    track_gap_range = config.get('params', 'track_gap_range')


    # 获取ID列表
    target_ids = fetch_target_ids(table, start_time, end_time, tests_sample_size)
    print(f"获取到的ID列表: {target_ids}")

    # 生成正样本
    positive_samples = extract_track_points(table, target_ids, end_time, track_extract_time, track_gap_range, track_sample_size)

    # # 将数据写入Excel
    write_to_excel(positive_samples, "trajectory_data.xlsx", "轨迹数据")
    generate_viewer_html()


def generate_viewer_html():
    """
    扫描 'tracks' 目录下的 .json 文件，并生成一个用于展示所有轨迹的 HTML 文件。
    """
    tracks_dir = 'tracks'
    html_template_path = 'web/main.html'
    output_html_path = 'web/view_all_tracks.html'
    
    # 1. 查找 'tracks' 目录下的所有 .json 文件
    try:
        # 使用 sorted() 确保文件顺序一致
        json_files = sorted([f for f in os.listdir(tracks_dir) if f.endswith('.json')])
        print(f"找到 {len(json_files)} 个轨迹文件。")
    except FileNotFoundError:
        print(f"错误: 找不到目录 '{tracks_dir}'。请确保该目录存在。")
        return

    # 2. 创建要注入的 JavaScript 代码
    # json.dumps 会正确处理任何特殊字符
    js_code = f'<script>var TRACK_FILES = {json.dumps(json_files)};</script>'
    
    # 3. 读取 HTML 模板文件
    try:
        with open(html_template_path, 'r', encoding='utf-8') as f:
            html_content = f.read()
    except FileNotFoundError:
        print(f"错误: 找不到模板文件 '{html_template_path}'。")
        return

    # 4. 将 JavaScript 代码注入到主脚本之前
    injection_marker = '<script src="scripts.js"></script>'
    if injection_marker not in html_content:
        print(f"错误: 在模板文件中找不到注入点 '{injection_marker}'。")
        return
        
    # 将标题修改为更合适的名称
    html_content = html_content.replace('<title>轨迹展示</title>', '<title>所有轨迹展示</title>')
    
    new_html_content = html_content.replace(
        injection_marker,
        f'    {js_code}\n    {injection_marker}'
    )
    
    # 5. 写入新的 HTML 文件
    with open(output_html_path, 'w', encoding='utf-8') as f:
        f.write(new_html_content)
        
    print(f"成功生成 '{output_html_path}'。")
    print("现在您可以在浏览器中直接打开此文件查看所有轨迹。")

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    main()

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
